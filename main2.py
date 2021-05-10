from logando import *
import getpass
import openpyxl


def executa(id_colecao, nome_arquivo, caminho='', alterar_recebimento=False):
    try:
        estoque = ShopifyApp()
        dataatual = estoque.dia_atual()
        comparando_datas = Comparacao()

        inventario = estoque.get_collection(id_colecao)

        cabecalho = (
        'Produto', '', '', '', 'Estoque Recebimento', '', '', 'Estoque Atual', '', '', 'Ultima verificacao', '',
        'Dias percorridos')

        arquivo_excel = openpyxl.Workbook()
        contador = 2

        if caminho != '':
            planilha_existente = openpyxl.load_workbook(filename=caminho)
            planilha1 = planilha_existente['Estoque']
            dataplanilha = planilha1['K2'].value
            diferenca = comparando_datas.compara_datas(dataatual, dataplanilha)

            if alterar_recebimento==True:
                for value in inventario:
                    inserir_tupla = (value['estoque'])
                    planilha1.cell(row=contador, column=5, value=inserir_tupla)
                    planilha1.cell(row=contador, column=8, value='')
                    planilha1.cell(row=contador, column=11, value=dataatual)
                    planilha1.cell(row=contador, column=13, value=diferenca)
                    contador = contador + 1
            else:
                for value in inventario:
                    inserir_tupla = (value['estoque'])
                    planilha1.cell(row=contador, column=8, value=inserir_tupla)
                    planilha1.cell(row=contador, column=11, value=dataatual)
                    planilha1.cell(row=contador, column=13, value=diferenca)
                    contador = contador + 1

            planilha_existente.save(caminho)

        else:
            planilha1 = arquivo_excel.active
            planilha1.title = 'Estoque'
            planilha1.append(cabecalho)
            for value in inventario:
                inserir_tupla = (value['nome'], '', '', '', value['estoque'], '', '', '', '', '', dataatual, '')
                planilha1.append(inserir_tupla)
            planilha1.cell(row=80, column=1, value=id_colecao)
            arquivo_excel.save(f'/Users/' + getpass.getuser() + f'/Desktop/' + f'{nome_arquivo}.xlsx')
    except Exception as error:
        print("Erro ao executar codigo: ", error)
